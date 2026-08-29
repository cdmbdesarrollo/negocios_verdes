# -*- coding: utf-8 -*-
"""
Genera 0025_datos_cdmb_negocios_verdes.sql a partir del Excel real de CDMB
(BASE_ACTUALIZADA_NV_ka.xlsx, hoja "BASE DE DATOS", 304 filas x 73
columnas). No se ejecuta como parte de la app — es una herramienta de una
sola vez, documentada acá para que quede claro cómo se produjo la
migración de datos. Reglas seguidas (acordadas con el usuario):

- Nunca se inventa un dato que el Excel no trae: si una celda está vacía o
  es un valor-sentinela ("NA", "N/A", "#N/A", "PENDIENTE"), el campo queda
  NULL en la base, salvo categoría oficial (ver más abajo, es NOT NULL).
- Categoría oficial es la única excepción: como negocios.categoria_oficial_id
  es NOT NULL, los negocios sin categoría reconocible quedan con la
  categoría-comodín "Pendiente de clasificar" (creada en este mismo script)
  en vez de con un texto inventado — así el admin los encuentra fácil y
  sabe que faltan por clasificar.
- Subcategoría/actividad productiva NO son obligatorias (son tablas puente
  opcionales) — si no hay match reconocible contra la taxonomía ya
  cargada, simplemente no se inserta esa fila puente, sin inventar ni
  usar comodín.
- Un negocio sin MUNICIPIO no se puede insertar (columna NOT NULL con
  CHECK de los 13 valores fijos) — esos quedan fuera del script, listados
  en el reporte final para completarlos a mano.
- activo siempre queda en FALSE al importar (ningún negocio del Excel
  trae foto de portada, y la tabla exige foto para poder marcarse activo)
  — CDMB decide caso por caso cuándo publicar cada uno, subiendo su foto
  desde /admin/negocios. novedad guarda el estado original de la base
  (ACTIVO/RETIRADO/SUSPENDIDO/...) para que el admin sepa cuáles corresponde
  publicar primero.

Uso: python generar_0025.py <ruta_excel> > 0025_datos_cdmb_negocios_verdes.sql
(el reporte de incidencias se imprime aparte, a un .txt, ver REPORTE_PATH).
"""
import openpyxl
import sys
import re
import uuid
import datetime

SQL_PATH = sys.argv[2] if len(sys.argv) > 2 else '0025_datos_cdmb_negocios_verdes.sql'
REPORTE_PATH = sys.argv[3] if len(sys.argv) > 3 else 'reporte_import_negocios.txt'

MUNICIPIOS_VALIDOS = [
    'Bucaramanga', 'Floridablanca', 'Girón', 'Piedecuesta', 'Vetas',
    'California', 'Suratá', 'Matanza', 'Charta', 'Tona', 'El Playón',
    'Rionegro', 'Lebrija',
]
MUNICIPIO_NORMALIZA = {'giron': 'Girón', 'floridablanca': 'Floridablanca'}

SENTINELAS = {
    'na', 'n/a', '#n/a', 'pendiente', 'ninguna', 'ninguno', '-', 'nulo',
    'no aplica', 'sin informacion', 'sin información',
}

CATEGORIA_MAP = {
    'bioproductos y servicios sostenibles': 'bioproductos-servicios-sostenibles',
    'ecoproductos industriales': 'ecoproductos-industriales',
    'ecoproductors industriales': 'ecoproductos-industriales',
    'productos por la calidad ambiental': 'calidad-ambiental',
}

SUBCATEGORIA_MAP = {
    'agrosistemas sostenibles': 'agrosistemas-sostenibles',
    'agroindustria sostenible': 'agroindustria-sostenible',
    'agroinsdustrial sostenible': 'agroindustria-sostenible',
    'biocomercio': 'biocomercio',
    'biotecnologia': 'biotecnologia',
    'biotecnología': 'biotecnologia',
    'turismo sostenible': 'turismo-sostenible',
    'aprovechamiento y valorización de residuos': 'aprovechamiento-valorizacion-residuos',
    'aprovechamiento y valoracion de residuos': 'aprovechamiento-valorizacion-residuos',
    'moda sostenible': 'moda-sostenible',
    'construcción e infraestructura sostenible': 'construccion-infraestructura-sostenible',
    'construccion e infraestructura sostenible': 'construccion-infraestructura-sostenible',
    'empaques y envases ecológicos': 'empaques-envases-ecologicos',
    'empaques y envases ecologicos': 'empaques-envases-ecologicos',
    'tecnologías verdes': 'tecnologias-verdes',
    'tecnologias verdes': 'tecnologias-verdes',
    'negocios asociados con la preservación y restauración  de ecosistemas': 'preservacion-restauracion-ecosistemas',
    'negocios asociados con la preservacion y restauracion de ecosistemas': 'preservacion-restauracion-ecosistemas',
    'transporte sostenible': 'transporte-sostenible',
}

ACTIVIDAD_MAP = {
    'agricultura orgánica': 'agricultura-organica',
    'agricultura organica': 'agricultura-organica',
    'agroecología': 'agroecologia',
    'agroecologia': 'agroecologia',
    'agricultura sostenible': 'agricultura-sostenible',
    'ganadería sostenible': 'ganaderia-sostenible',
    'ganaderia sostenible': 'ganaderia-sostenible',
    'acuicultura y pesca sostenible': 'acuicultura-pesca-sostenible',
    'agroindustrial alimentario': 'agroindustrial-alimentario',
    'agroindustrial no alimentario': 'agroindustrial-no-alimentario',
    'recursos genéticos y productos derivados': 'recursos-geneticos-productos-derivados',
    'recursos geneticos y productos derivados': 'recursos-geneticos-productos-derivados',
    'productos derivados de la fauna silvestre': 'productos-fauna-silvestre',
    'no maderables': 'no-maderables',
    'maderables': 'maderables',
    'productos de la biotecnología': 'productos-biotecnologia',
    'productos de la biotecnologia': 'productos-biotecnologia',
    'porductos de la biotecnología': 'productos-biotecnologia',
    'servicios de turismo de naturaleza': 'servicios-turismo-naturaleza',
    'otros servicios de turismo sostenible': 'otros-servicios-turismo-sostenible',
    'aprovechamiento de residuos orgánicos': 'aprovechamiento-residuos-organicos',
    'aprovechamiento de residuos organicos': 'aprovechamiento-residuos-organicos',
    'aprovechamiento de residuos inorgánicos': 'aprovechamiento-residuos-inorganicos',
    'aprovechamiento de residuos inorganicos': 'aprovechamiento-residuos-inorganicos',
    'textiles sostenibles': 'textiles-sostenibles',
    'confección y manufactura': 'confeccion-manufactura',
    'confeccion y manufactura': 'confeccion-manufactura',
    'joyería, artesanía y bisutería': 'joyeria-artesania-bisuteria',
    'joyeria, artesania y bisuteria': 'joyeria-artesania-bisuteria',
    'construcción de edificaciones e infraestructura sostenible': 'construccion-edificaciones-infraestructura',
    'biomateriales, eco materiales y equipos ecoficientes': 'biomateriales-ecomateriales-equipos-ecoeficientes',
    'biomateriales, eco materiales y equipos ecoeficientes': 'biomateriales-ecomateriales-equipos-ecoeficientes',
    'biopolímeros, fibras naturales, empaques y envases reciclabes': 'biopolimeros-fibras-empaques-reciclables',
    'biopolimeros, fibras naturales, empaques y envases reciclables': 'biopolimeros-fibras-empaques-reciclables',
    'generación y/o comercialización de energía a partir de fncer': 'generacion-comercializacion-energia-fncer',
    'generacion y/o comercializacion de energia a partir de fncer': 'generacion-comercializacion-energia-fncer',
    'tecnologias de información ambiental y otras tecnologías limpias': 'tecnologias-informacion-ambiental',
    'tecnologias de informacion ambiental y otras tecnologias limpias': 'tecnologias-informacion-ambiental',
    'preservación': 'preservacion',
    'preservacion': 'preservacion',
    'restauración': 'restauracion',
    'restauracion': 'restauracion',
    'recuperación y remediación': 'recuperacion-remediacion',
    'recuperacion y remediacion': 'recuperacion-remediacion',
    'motorizado': 'motorizado',
    'no motorizado': 'no-motorizado',
}


def colidx(headers, name):
    """Exact match primero (evita falsos positivos como 'ICA' encontrando
    'NATURAL - JURÍDICA' por contener 'ica' como substring), substring como
    respaldo para los términos deliberadamente parciales (ej. 'TIEMPO DE
    CONSTITUCIÓN' calzando con el encabezado completo más largo)."""
    target = name.strip().lower()
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == target:
            return i
    for i, h in enumerate(headers):
        if h and target in str(h).strip().lower():
            return i
    return None


def limpio(v):
    """None si la celda está vacía o es un valor-sentinela; si no, el
    string ya recortado. Nunca inventa contenido."""
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.lower() in SENTINELAS:
        return None
    return s


def sql_str(v):
    if v is None:
        return 'null'
    s = str(v).replace("'", "''")
    return f"'{s}'"


def sql_bool(v):
    return 'true' if v else 'false'


def sql_int(v):
    return 'null' if v is None else str(int(v))


def sql_num(v):
    return 'null' if v is None else str(float(v))


def sql_date(v):
    if v is None:
        return 'null'
    if isinstance(v, (datetime.date, datetime.datetime)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return 'null'


def telefono_str(v):
    s = limpio(v)
    if s is None:
        return None
    if isinstance(v, (int, float)):
        return str(int(v))
    return s


def slugificar(texto):
    s = texto.strip().lower()
    s = (s.replace('á', 'a').replace('é', 'e').replace('í', 'i')
           .replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n'))
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s or 'vereda'


def normalizar_vereda(texto):
    """Title case simple, colapsa espacios — la base trae mayúsculas
    sueltas ('BARRIO BLANCO'), minúsculas sueltas ('la union') y mixtas
    para la MISMA vereda; normalizar la escritura no es inventar el dato,
    solo lo hace un catálogo consistente."""
    partes = re.sub(r'\s+', ' ', texto.strip()).split(' ')
    return ' '.join(p if len(p) <= 2 else p.capitalize() for p in partes)


def normalizar_novedad(texto):
    """Unifica los 3 valores de NOVEDAD que en la práctica son lo mismo
    (pedido explícito): 'RETIRADO', 'ACTIVO (RETIRADO)' y 'ACTIVO
    (RETIRADO) P' (la "P" suelta es una anotación, no un estado distinto)
    quedan todos como 'RETIRADO' — ninguno de los 3 es 'ACTIVO' a secas de
    todas formas, así que esto no cambia qué negocios se publican, solo
    limpia el texto que ve el admin. SUSPENDIDO/ACTIVO (SUSPENDIDO) no se
    tocan, no fueron parte del pedido."""
    if texto is None:
        return None
    v = texto.strip()
    if 'RETIRADO' in v.upper():
        return 'RETIRADO'
    return v


DMS_RE = re.compile(
    r"^\(?-?\)?\s*(\d{1,3})[°:]\s*(\d{1,2})['’:]\s*([\d.,]+)")


def parse_dms(texto):
    """Devuelve grados decimales (siempre positivos) o None si el string
    no calza con un DMS válido (minutos/segundos fuera de rango, formato
    irreconocible, etc.) — nunca se adivina un valor a medias."""
    if not texto:
        return None
    s = str(texto).strip()
    if s.lower() in SENTINELAS or s == '':
        return None
    # Decimal ya listo (ej. 72.8746247) — algunas filas ya traen así.
    try:
        val = float(s.replace(',', '.'))
        if 0 < val < 200:
            return abs(val)
    except ValueError:
        pass
    m = DMS_RE.match(s)
    if not m:
        return None
    grados, minutos, segundos = m.groups()
    try:
        grados, minutos = int(grados), int(minutos)
        segundos = float(segundos.replace(',', '.'))
    except ValueError:
        return None
    if minutos >= 60 or segundos >= 60:
        return None
    return grados + minutos / 60 + segundos / 3600


def main():
    ruta = sys.argv[1]
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def col(name):
        return colidx(headers, name)

    i = {
        'nvemp': col('NV / EMP'), 'anio': col('AÑO'), 'novedad': col('NOVEDAD'),
        'municipio': col('MUNICIPIO'), 'vereda': col('VEREDA'),
        'razon': col('RAZÓN SOCIAL'), 'tiempo_const': col('TIEMPO DE CONSTITUCIÓN'),
        'nit': col('CC / NIT'), 'naturaleza': col('NATURAL - JURÍDICA'),
        'rut': col('RUT - CÁMARA'), 'descripcion': col('DESCRIPCION'),
        'producto': col('PRODUCTO'), 'actividad': col('ACTIVIDAD PRODUCTIVA'),
        'categoria': col('CATEGORÍA'), 'subcategoria': col('SUBCATEGORIA'),
        'correo': col('CORREO'), 'repleg': col('REPRESENTANTE LEGAL'),
        'delegado': col('DELEGADO'), 'telefono': col('TELÉFONO'),
        'direccion': col('DIRECCIÓN'), 'resp_cdmb': col('RESPONSABLE CDMB'),
        'rnt': col('REGISTRO NACIONAL DE TURISMO'), 'uso_suelo': col('USO DEL SUELO'),
        'conc_aguas': col('CONCESIÓN DE AGUAS'),
        'vertimientos': col('VERTIMIENTOS'),
        'pueaa': col('PUEAA'), 'pgris': col('PGRIS'),
        'pozo': col('POZO SÉPTICOS'), 'alcantarillado': col('ALCANTARILLADO'),
        'ica': col('ICA (REGISTRO'), 'invima': col('INVIMA'),
        'cert_animales': col('CERTIFICADO DE TENENCIA'),
        'bpa': col('BUENAS PRÁCTICAS AGRICOLAS'),
        'bpa_apicola': col('BUENAS PRÁCTICAS APICOLAS'),
        'reg_apicola': col('REGISTRO APICOLA'),
        'interv_cauce': col('INTERVENCIÓN DE CAUCE'),
        'capacidad_carga': col('CAPACIDAD DE CARGA'),
        'sstt': col('SSTT'), 'canal_venta': col('B2B'),
        'exportacion': col('EXPORTACION'), 'huella': col('HUELLA DE CARBONO'),
        'fort_tec': col('FORTALECIMIENTO TÉCNICO'),
        'fort_aca': col('FORTALECIMIENTO ACADEMICO'),
        'fort_fin': col('FORTALECIMIENTO FINANCIERO'),
        'internac': col('INTERNACIONALIZACIÓN'),
        'certificaciones': col('CERTIFICACIONES'),
        'posic_marca': col('POSICIONAMIENTO DE MARCA'),
        'fort_amb': col('FORTALEZAS - AMBIENTAL'),
        'fort_soc': col('FORTALEZAS - SOCIAL'),
        'fort_eco': col('FORTALEZAS - ECONÓMICO'),
        'deb_amb': col('DEBILIDADES - AMBIENTAL'),
        'deb_soc': col('DEBILIDADES - SOCIAL'),
        'deb_fin': col('DEBILIDADES - FINANCIERA'),
        'beneficios': col('BENEFICIOS RECIBIDOS'),
        'aval': col('AVAL'), 'sello': col('SELLO MARCA'),
        'codigo_marca': col('CODIGO MARCA'), 'tipo_nv': col('TIPO DE NEGOCIO VERDE'),
        'aplicacion_2025': col('APLICACIÓN DE FICHA 2025'),
        'observaciones': col('OBSERVACIONES'),
        'este': col('ESTE'), 'norte': col('NORTE'), 'cota': col('COTA'),
    }
    # Las 2 columnas "FECHA DE VENCIMIENTO" repetidas: la primera ocurrencia
    # sigue a CONCESIÓN DE AGUAS, la segunda a VERTIMIENTOS, la tercera a
    # ICA, la cuarta a INVIMA — se ubican por posición relativa, no por
    # nombre (el nombre se repite 4 veces).
    venc_idxs = [n for n, h in enumerate(headers)
                 if h and 'fecha de vencimiento' in str(h).strip().lower()]
    i['conc_aguas_venc'], i['vertimientos_venc'], i['ica_venc'], i['invima_venc'] = (
        venc_idxs + [None, None, None, None])[:4]

    # Preámbulo: idempotente a propósito (WHERE NOT EXISTS / ON CONFLICT DO
    # NOTHING en los 3 statements) para poder repetirlo al principio de
    # CADA archivo partido (ver chunking al final) sin duplicar nada si el
    # admin corre los archivos en cualquier orden o repite uno.
    preambulo = []
    preambulo.append(
        "-- Categoría comodín para negocios sin categoría reconocible en el\n"
        "-- Excel (columna NOT NULL, no se puede dejar en blanco). activo=false\n"
        "-- así no aparece como chip de filtro en el buscador público.\n"
        "insert into categorias_oficiales (nombre, slug, descripcion, icono, orden, activo)\n"
        "select 'Pendiente de clasificar', 'pendiente-clasificar',\n"
        "  'Negocio importado de la base CDMB sin categoría oficial asignada todavía — revisar y corregir desde /admin/negocios.',\n"
        "  '⏳', 99, false\n"
        "where not exists (select 1 from categorias_oficiales where slug = 'pendiente-clasificar');"
    )
    # Único negocio de prueba en producción — confirmado con el usuario.
    # DELETE sin WHERE de un id inexistente no es un error, así que también
    # es seguro repetirlo.
    preambulo.append("delete from negocios where nombre = 'Bucarretes SAS BIC';")

    # --- Veredas ---------------------------------------------------------
    veredas_vistas = {}  # (municipio, slug) -> nombre normalizado
    for r in rows:
        municipio_raw = limpio(r[i['municipio']])
        vereda_raw = limpio(r[i['vereda']])
        if not municipio_raw or not vereda_raw:
            continue
        municipio = MUNICIPIO_NORMALIZA.get(municipio_raw.lower(), municipio_raw)
        if municipio not in MUNICIPIOS_VALIDOS:
            continue
        nombre = normalizar_vereda(vereda_raw)
        slug = slugificar(nombre)
        veredas_vistas.setdefault((municipio, slug), nombre)

    if veredas_vistas:
        valores = [
            f"\n  ({sql_str(mun)}, {sql_str(nom)}, {sql_str(slug)})"
            for (mun, slug), nom in sorted(veredas_vistas.items())
        ]
        preambulo.append(
            '-- Veredas encontradas en la base real (normalizadas de escritura).\n'
            'insert into veredas (municipio, nombre, slug) values'
            + ','.join(valores)
            + '\non conflict (municipio, slug) do nothing;'
        )

    # --- Negocios ----------------------------------------------------------
    reporte = {
        'sin_municipio': [], 'categoria_pendiente': [], 'sin_coordenadas': [],
        'coordenadas_no_parseables': [], 'total_importados': 0,
        'por_novedad': {}, 'badges': {
            'emprendimiento_verde': 0, 'sello_marca': 0, 'avalado': 0,
            'sin_senal_explicita_fallback_emprendimiento': 0,
        },
    }

    # Un bloque de texto por negocio (su INSERT + sus puentes + sus
    # puntajes juntos) — así el chunking de más abajo puede cortar entre
    # negocios sin nunca partir un negocio a la mitad ni depender de que
    # otro archivo ya haya corrido antes (cada negocio es autocontenido
    # salvo el preámbulo, que se repite en cada archivo).
    bloques_negocio = []

    anios_puntaje = [
        ('PUNTAJE 2020', 2020), ('PUNTAJE 2021', 2021), ('PUNTAJE 2022', 2022),
        ('PUNTAJE 2023', 2023), ('PUNTAJE 2024', 2024), ('PUNTAJE 2025', 2025),
    ]
    idx_puntajes = [(col(nombre), anio) for nombre, anio in anios_puntaje]

    for r in rows:
        nombre = limpio(r[i['razon']])
        if not nombre:
            continue  # no debería pasar (0 nulos verificado), guarda igual

        municipio_raw = limpio(r[i['municipio']])
        if not municipio_raw:
            reporte['sin_municipio'].append(nombre)
            continue
        municipio = MUNICIPIO_NORMALIZA.get(municipio_raw.lower(), municipio_raw)
        if municipio not in MUNICIPIOS_VALIDOS:
            reporte['sin_municipio'].append(f'{nombre} (municipio irreconocible: {municipio_raw})')
            continue

        negocio_id = str(uuid.uuid4())

        vereda_raw = limpio(r[i['vereda']])
        vereda_sql = 'null'
        if vereda_raw:
            slug = slugificar(normalizar_vereda(vereda_raw))
            vereda_sql = (f"(select id from veredas where municipio = {sql_str(municipio)} "
                          f"and slug = {sql_str(slug)})")

        # --- categoría / subcategoría / actividad ---
        categoria_raw = limpio(r[i['categoria']])
        categoria_slug = (CATEGORIA_MAP.get(categoria_raw.lower())
                           if categoria_raw else None)
        if categoria_slug is None:
            categoria_slug = 'pendiente-clasificar'
            reporte['categoria_pendiente'].append(nombre)
        categoria_sql = f"(select id from categorias_oficiales where slug = {sql_str(categoria_slug)})"

        subcategoria_raw = limpio(r[i['subcategoria']])
        subcategoria_slug = (SUBCATEGORIA_MAP.get(subcategoria_raw.lower())
                              if subcategoria_raw else None)

        actividad_raw = limpio(r[i['actividad']])
        actividad_slug = (ACTIVIDAD_MAP.get(actividad_raw.lower())
                           if actividad_raw else None)

        # --- coordenadas ---
        este = parse_dms(r[i['este']])
        norte = parse_dms(r[i['norte']])
        if este is None and norte is None:
            if limpio(r[i['este']]) is None and limpio(r[i['norte']]) is None:
                reporte['sin_coordenadas'].append(nombre)
            else:
                reporte['coordenadas_no_parseables'].append(
                    f"{nombre} (ESTE={r[i['este']]!r} NORTE={r[i['norte']]!r})")
        longitud = -este if este is not None else None  # Santander es oeste
        latitud = norte if norte is not None else None
        # Crudos, tal cual venían en el Excel — para poder corregir/recalcular
        # a mano desde el admin si la conversión automática no da bien.
        este_raw = limpio(r[i['este']])
        norte_raw = limpio(r[i['norte']])

        # --- badges (3 independientes, ver 0021) ---
        nvemp = (limpio(r[i['nvemp']]) or '').upper()
        aval_raw = (limpio(r[i['aval']]) or '').upper()
        sello_raw = (limpio(r[i['sello']]) or '').upper()
        emprendimiento_verde = nvemp in ('EMPRENDIMIENTO VERDE', 'EMPRENDIMIENTO')
        sello_marca = sello_raw == 'SI' or nvemp == 'NEGOCIO VERDE - SELLO MARCA'
        avalado = aval_raw == 'SI' or nvemp == 'NEGOCIO VERDE AVALADO'
        # Pedido explícito de CDMB: TODO negocio real de su base tiene
        # siempre alguna de las 3 — "Emprendimiento Verde" es la categoría
        # base/de entrada al programa (nadie llega directo a Sello Marca o
        # Avalado sin pasar por ahí), así que cuando ninguna de las 3
        # señales de arriba prendió (67 casos de 'NEGOCIO VERDE' a secas
        # sin AVAL/SELLO MARCA en 'SI'), el negocio de todas formas ES un
        # Emprendimiento Verde — no queda sin clasificar.
        if not (emprendimiento_verde or sello_marca or avalado):
            emprendimiento_verde = True
            reporte['badges']['sin_senal_explicita_fallback_emprendimiento'] += 1
        if emprendimiento_verde:
            reporte['badges']['emprendimiento_verde'] += 1
        if sello_marca:
            reporte['badges']['sello_marca'] += 1
        if avalado:
            reporte['badges']['avalado'] += 1

        # --- naturaleza jurídica ---
        nat_raw = (limpio(r[i['naturaleza']]) or '').lower()
        naturaleza = None
        if nat_raw.startswith('natural'):
            naturaleza = 'Natural'
        elif nat_raw.startswith('jur'):
            naturaleza = 'Jurídica'

        novedad = normalizar_novedad(limpio(r[i['novedad']]))
        reporte['por_novedad'][novedad] = reporte['por_novedad'].get(novedad, 0) + 1
        # activo=true directo para todo lo que CDMB ya marca ACTIVO — la
        # foto de portada dejó de ser obligatoria (0023_foto_portada_opcional.sql),
        # así que ya no hay motivo para dejarlos todos ocultos a esperar
        # que alguien suba una foto uno por uno.
        activo = novedad == 'ACTIVO'

        # Repetido a propósito en telefono Y whatsapp (pedido explícito):
        # casi todos los números de la base son celular, no fijo — así
        # funcionan tanto "Llamar" como el botón de WhatsApp sin depender
        # de que alguien lo complete después. Normaliza a formato wa.me
        # (57 + 10 dígitos) solo cuando el número ya parece un celular
        # colombiano de 10 dígitos empezando en 3 — cualquier otro formato
        # se deja tal cual, sin inventar un indicativo que no se sabe si
        # aplica.
        telefono_valor = telefono_str(r[i['telefono']])
        whatsapp_valor = telefono_valor
        if whatsapp_valor and len(whatsapp_valor) == 10 and whatsapp_valor.startswith('3'):
            whatsapp_valor = '57' + whatsapp_valor

        descripcion = limpio(r[i['descripcion']])
        descripcion_corta = None
        if descripcion:
            if len(descripcion) <= 157:
                descripcion_corta = descripcion
            else:
                corte = descripcion[:157].rsplit(' ', 1)[0]
                descripcion_corta = corte + '…'

        campos = {
            'id': sql_str(negocio_id),
            'nombre': sql_str(nombre),
            'slug': f"generar_slug_unico({sql_str(nombre)}, {sql_str(negocio_id)})",
            'categoria_oficial_id': categoria_sql,
            'municipio': sql_str(municipio),
            'vereda_id': vereda_sql,
            'direccion': sql_str(limpio(r[i['direccion']])),
            'latitud': sql_num(latitud),
            'longitud': sql_num(longitud),
            'descripcion_corta': sql_str(descripcion_corta),
            'descripcion': sql_str(descripcion),
            'producto': sql_str(limpio(r[i['producto']])),
            'telefono': sql_str(telefono_valor),
            'whatsapp': sql_str(whatsapp_valor),
            'email': sql_str(limpio(r[i['correo']])),
            'representante_legal': sql_str(limpio(r[i['repleg']])),
            'nit': sql_str(limpio(r[i['nit']])),
            'naturaleza_juridica': sql_str(naturaleza),
            'delegado': sql_str(limpio(r[i['delegado']])),
            'tiempo_constitucion': sql_str(limpio(r[i['tiempo_const']])),
            'rut_camara_comercio': sql_str(limpio(r[i['rut']])),
            'responsable_cdmb': sql_str(limpio(r[i['resp_cdmb']])),
            'novedad': sql_str(novedad),
            'tipo_negocio_verde': sql_str(limpio(r[i['tipo_nv']])),
            'codigo_marca': sql_str(limpio(r[i['codigo_marca']])),
            'anio_registro': sql_int(r[i['anio']] if isinstance(r[i['anio']], (int, float)) else None),
            'cota_msnm': sql_str(limpio(r[i['cota']])),
            'este': sql_str(este_raw),
            'norte': sql_str(norte_raw),
            'aplicacion_ficha_2025': sql_str(limpio(r[i['aplicacion_2025']])),
            'observaciones': sql_str(limpio(r[i['observaciones']])),
            'registro_nacional_turismo': sql_str(limpio(r[i['rnt']])),
            'uso_suelo': sql_str(limpio(r[i['uso_suelo']])),
            'concesion_aguas': sql_str(limpio(r[i['conc_aguas']])),
            'concesion_aguas_vencimiento': sql_date(r[i['conc_aguas_venc']]) if i['conc_aguas_venc'] is not None else 'null',
            'vertimientos': sql_str(limpio(r[i['vertimientos']])),
            'vertimientos_vencimiento': sql_date(r[i['vertimientos_venc']]) if i['vertimientos_venc'] is not None else 'null',
            'pueaa': sql_str(limpio(r[i['pueaa']])),
            'pgris': sql_str(limpio(r[i['pgris']])),
            'pozo_septico': sql_str(limpio(r[i['pozo']])),
            'alcantarillado': sql_str(limpio(r[i['alcantarillado']])),
            'ica': sql_str(limpio(r[i['ica']])),
            'ica_vencimiento': sql_date(r[i['ica_venc']]) if i['ica_venc'] is not None else 'null',
            'invima': sql_str(limpio(r[i['invima']])),
            'invima_vencimiento': sql_date(r[i['invima_venc']]) if i['invima_venc'] is not None else 'null',
            'certificado_tenencia_animales': sql_str(limpio(r[i['cert_animales']])),
            'buenas_practicas_agricolas': sql_str(limpio(r[i['bpa']])),
            'buenas_practicas_apicolas': sql_str(limpio(r[i['bpa_apicola']])),
            'registro_apicola': sql_str(limpio(r[i['reg_apicola']])),
            'intervencion_cauce': sql_str(limpio(r[i['interv_cauce']])),
            'capacidad_carga': sql_str(limpio(r[i['capacidad_carga']])),
            'sstt': sql_str(limpio(r[i['sstt']])),
            'canal_venta': sql_str(limpio(r[i['canal_venta']])),
            'exportacion': sql_str(limpio(r[i['exportacion']])),
            'huella_carbono': sql_str(limpio(r[i['huella']])),
            'fortalecimiento_tecnico': sql_str(limpio(r[i['fort_tec']])),
            'fortalecimiento_academico': sql_str(limpio(r[i['fort_aca']])),
            'fortalecimiento_financiero': sql_str(limpio(r[i['fort_fin']])),
            'internacionalizacion': sql_str(limpio(r[i['internac']])),
            'certificaciones': sql_str(limpio(r[i['certificaciones']])),
            'posicionamiento_marca': sql_str(limpio(r[i['posic_marca']])),
            'beneficios_ventanilla': sql_str(limpio(r[i['beneficios']])),
            'fortalezas_ambiental': sql_str(limpio(r[i['fort_amb']])),
            'fortalezas_social': sql_str(limpio(r[i['fort_soc']])),
            'fortalezas_economico': sql_str(limpio(r[i['fort_eco']])),
            'debilidades_ambiental': sql_str(limpio(r[i['deb_amb']])),
            'debilidades_social': sql_str(limpio(r[i['deb_soc']])),
            'debilidades_financiera': sql_str(limpio(r[i['deb_fin']])),
            'emprendimiento_verde': sql_bool(emprendimiento_verde),
            'sello_marca': sql_bool(sello_marca),
            'avalado': sql_bool(avalado),
            'destacado': 'false',
            'activo': sql_bool(activo),
        }

        columnas = ', '.join(campos)
        valores = ', '.join(campos.values())
        lineas = [f"insert into negocios ({columnas}) values ({valores});"]

        lineas.append(
            f"insert into negocios_categorias (negocio_id, categoria_oficial_id) "
            f"values ({sql_str(negocio_id)}, {categoria_sql});")
        if subcategoria_slug:
            lineas.append(
                f"insert into negocios_subcategorias (negocio_id, subcategoria_id) "
                f"select {sql_str(negocio_id)}, id from subcategorias where slug = {sql_str(subcategoria_slug)};")
        if actividad_slug:
            lineas.append(
                f"insert into negocios_actividades (negocio_id, actividad_productiva_id) "
                f"select {sql_str(negocio_id)}, id from actividades_productivas where slug = {sql_str(actividad_slug)};")

        for idx_col, anio in idx_puntajes:
            valor = r[idx_col] if idx_col is not None else None
            if isinstance(valor, (int, float)):
                lineas.append(
                    f"insert into negocio_puntajes (negocio_id, anio, puntaje) "
                    f"values ({sql_str(negocio_id)}, {anio}, {float(valor)}) "
                    f"on conflict (negocio_id, anio) do nothing;")

        bloques_negocio.append(f"-- {nombre}\n" + '\n'.join(lineas))
        reporte['total_importados'] += 1

    # --- Partir en varios archivos chicos -----------------------------
    # El editor SQL del dashboard de Supabase truncó el archivo único de
    # ~1 MB a mitad de una línea (reportado por el usuario: "unterminated
    # quoted string") — es una limitación del editor web, no un error de
    # sintaxis real. Se parte en archivos de ~120 KB (bien por debajo de
    # donde truncó el de 1 MB) cortando siempre entre negocios completos,
    # nunca a la mitad de uno. Cada archivo es una transacción propia
    # (begin/commit) con el preámbulo repetido (es idempotente) al
    # principio, así se pueden correr en cualquier orden, o repetir uno
    # sin duplicar nada, sin depender de que otro archivo ya haya corrido.
    LIMITE_BYTES = 120_000
    preambulo_txt = '\n\n'.join(preambulo)
    encabezado_comun = (
        "-- Generado por lib/migraciones/generar_0025.py desde "
        "BASE_ACTUALIZADA_NV_ka.xlsx — no editar a mano.\n"
        "-- Uno de varios archivos partidos (ver README.md) — correr TODOS, "
        "en cualquier orden, cada uno es su propia transacción.\n"
    )

    archivos = []
    bloque_actual = []
    tamano_actual = len(preambulo_txt.encode('utf-8'))
    for bloque in bloques_negocio:
        tamano_bloque = len(bloque.encode('utf-8'))
        if bloque_actual and tamano_actual + tamano_bloque > LIMITE_BYTES:
            archivos.append(bloque_actual)
            bloque_actual = []
            tamano_actual = len(preambulo_txt.encode('utf-8'))
        bloque_actual.append(bloque)
        tamano_actual += tamano_bloque
    if bloque_actual:
        archivos.append(bloque_actual)

    base_path = SQL_PATH[:-4] if SQL_PATH.endswith('.sql') else SQL_PATH
    total_partes = len(archivos)
    for idx, bloques in enumerate(archivos, start=1):
        partes = [
            'begin;\n',
            encabezado_comun,
            f'-- Parte {idx} de {total_partes}.\n',
            preambulo_txt,
            '\n\n-- Negocios de esta parte (INSERT + categoría/subcategoría/'
            'actividad + puntajes de cada uno, juntos).\n',
            '\n\n'.join(bloques),
            '\n\ncommit;\n',
        ]
        ruta_parte = f'{base_path}_{idx:02d}.sql'
        with open(ruta_parte, 'w', encoding='utf-8') as f:
            f.write('\n'.join(partes))

    with open(REPORTE_PATH, 'w', encoding='utf-8') as f:
        f.write(f"Generado en {total_partes} archivos: "
                f"{base_path}_01.sql .. {base_path}_{total_partes:02d}.sql\n\n")
        f.write(f"Total importados: {reporte['total_importados']} / {len(rows)}\n\n")
        f.write(f"Sin municipio ({len(reporte['sin_municipio'])}) — no se importaron, agregar a mano:\n")
        for n in reporte['sin_municipio']:
            f.write(f"  - {n}\n")
        f.write(f"\nCon categoría 'Pendiente de clasificar' ({len(reporte['categoria_pendiente'])}):\n")
        for n in reporte['categoria_pendiente']:
            f.write(f"  - {n}\n")
        f.write(f"\nSin ninguna coordenada en el Excel ({len(reporte['sin_coordenadas'])})\n")
        f.write(f"\nCoordenadas con formato no reconocible ({len(reporte['coordenadas_no_parseables'])}):\n")
        for n in reporte['coordenadas_no_parseables']:
            f.write(f"  - {n}\n")
        f.write(f"\nPor NOVEDAD original:\n")
        for k, v in sorted(reporte['por_novedad'].items(), key=lambda x: -x[1]):
            f.write(f"  {k}: {v}\n")
        f.write(f"\nBadges asignados: {reporte['badges']}\n")


if __name__ == '__main__':
    main()
